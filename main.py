import asyncio
import os
import json
import glob
import logging
import re
import datetime
from datetime import datetime, timedelta
from email.message import EmailMessage
import smtplib
import ssl
import pandas as pd
import openpyxl

# Viam imports
from viam.rpc.dial import DialOptions, Credentials
from viam.app.viam_client import ViamClient
from viam.app.data_client import DataClient
from viam.proto.app.data import Filter, BinaryID
from viam.proto.app import Location
from viam.utils import create_filter

##############################################################################
#                          Logging Configuration                             #
##############################################################################

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)

##############################################################################
#                           Helper Functions                                 #
##############################################################################

def load_credentials(creds_path: str = "creds.json") -> dict:
    logger.info(f"Loading credentials from {creds_path}")
    with open(creds_path, "r") as f:
        return json.load(f)

def build_dial_options(api_key: str, api_key_id: str) -> DialOptions:
    logger.info("Building DialOptions for Viam connection")
    return DialOptions(
        credentials=Credentials(type="api-key", payload=api_key),
        auth_entity=api_key_id
    )

async def connect_to_viam(api_key: str, api_key_id: str) -> ViamClient:
    logger.info("Connecting to Viam Cloud ...")
    dial_opts = build_dial_options(api_key, api_key_id)
    client = await ViamClient.create_from_dial_options(dial_opts)
    logger.info("Connected to Viam successfully.")
    return client

def build_time_filter(loc_ids: list[str], days_back: int = 1) -> Filter:
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(days=days_back)
    logger.info(f"Building time filter from {start_time} to {end_time} for loc_ids: {loc_ids}")
    return create_filter(
        location_ids=loc_ids,
        start_time=start_time,
        end_time=end_time
    )

async def fetch_location_names(cloud_client, loc_ids: list[str]) -> dict:
    loc_names = {}
    for loc_id in loc_ids:
        location: Location = await cloud_client.get_location(location_id=loc_id)
        loc_names[loc_id] = location.name
        logger.info(f"Fetched location '{location.name}' for ID '{loc_id}'")
    return loc_names

async def fetch_csv_metadata(
    data_client: DataClient,
    filter_obj: Filter,
    substring: str = "tracked-objects.csv"
) -> list:
    logger.info(f"Fetching metadata from Viam Cloud for files containing '{substring}'")
    csv_metadata = []
    last = None
    total_found = 0
    logger.info("Fetching metadata in batches ...")
    while True:
        logger.info("Fetching next batch of metadata ...")
        metadata_list, total_count, last = await data_client.binary_data_by_filter(
            filter_obj,
            limit=50,
            last=last,
            include_binary_data=False
        )
        if not metadata_list:
            break
        for item in metadata_list:
            if substring in item.metadata.file_name:
                csv_metadata.append(item)
        total_found += len(metadata_list)
        if last is None:
            break
    logger.info(
        f"Fetched a total of {total_found} metadata records (all file types), "
        f"with {len(csv_metadata)} matching '{substring}'."
    )
    return csv_metadata

async def download_csv_file(
    data_client: DataClient,
    meta_item,
    download_dir: str,
    loc_names: dict
) -> str:
    original_path = meta_item.metadata.file_name
    original_basename = os.path.basename(original_path)
    loc_id = meta_item.metadata.capture_metadata.location_id
    location_name = loc_names.get(loc_id, "UnknownLocation")
    new_filename = f"{location_name}-{original_basename}"
    local_path = os.path.join(download_dir, new_filename)
    if os.path.exists(local_path):
        logger.info(f"File already exists locally, skipping download: {local_path}")
        return local_path
    logger.info(f"Downloading file '{original_basename}' from Viam Cloud for location '{location_name}'...")
    from viam.proto.app.data import BinaryID
    b_id = BinaryID(
        file_id=meta_item.metadata.id,
        organization_id=meta_item.metadata.capture_metadata.organization_id,
        location_id=loc_id
    )
    data_list = await data_client.binary_data_by_ids([b_id])
    if not data_list:
        logger.warning(f"No actual data found for file {original_basename}")
        return None
    raw_bytes = data_list[0].binary
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    with open(local_path, "wb") as f:
        f.write(raw_bytes)
    logger.info(f"Downloaded CSV to {local_path}")
    return local_path

async def download_csvs_to_directory(
    data_client: DataClient,
    loc_ids: list[str],
    loc_names: dict,
    days_back: int,
    download_dir: str,
    csv_substring: str = "tracked-objects.csv"
) -> list[str]:
    downloaded_paths = []
    for loc_id in loc_ids:
        filter_obj = build_time_filter([loc_id], days_back)
        csv_metas = await fetch_csv_metadata(data_client, filter_obj, substring=csv_substring)
        logger.info(f"For location {loc_id}: Found {len(csv_metas)} metadata records for CSV matching '{csv_substring}'.")
        if not csv_metas:
            logger.info(f"No CSV files to download for location {loc_id}.")
            continue
        csv_metas.sort(key=lambda m: m.metadata.time_received, reverse=True)
        for meta_item in csv_metas:
            path = await download_csv_file(data_client, meta_item, download_dir, loc_names)
            if path:
                downloaded_paths.append(path)
    logger.info(f"Total new or verified CSV files downloaded: {len(downloaded_paths)}")
    return downloaded_paths

def clear_directory(directory: str) -> None:
    files = glob.glob(os.path.join(directory, "*"))
    for file in files:
        os.remove(file)
    logger.info(f"Cleared all files in directory: {directory}")

def gather_all_files_in_directory(directory: str) -> list[str]:
    pattern = os.path.join(directory, "*")
    files = [f for f in glob.glob(pattern) if os.path.isfile(f)]
    logger.info(f"Found {len(files)} files in directory: {directory}")
    return files

def send_email_with_attachments(
    smtp_host: str,
    smtp_port: int,
    sender_email: str,
    sender_password: str,
    recipient_email,  # can be a list or string
    subject: str,
    body: str,
    attachment_paths: list[str]
):
    logger.info(f"Preparing to send email to {recipient_email} with {len(attachment_paths)} attachment(s).")
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender_email
    if isinstance(recipient_email, list):
        msg["To"] = ", ".join(recipient_email)
    else:
        msg["To"] = recipient_email
    msg.set_content(body)
    for path in attachment_paths:
        if not os.path.exists(path):
            logger.warning(f"File not found, skipping attachment: {path}")
            continue
        with open(path, "rb") as f:
            file_data = f.read()
        filename = os.path.basename(path)
        if filename.lower().endswith(".csv"):
            maintype, subtype = "text", "csv"
        elif filename.lower().endswith(".pdf"):
            maintype, subtype = "application", "pdf"
        else:
            maintype, subtype = "application", "octet-stream"
        msg.add_attachment(file_data, maintype=maintype, subtype=subtype, filename=filename)
        logger.info(f"Attached file: {filename}")
    context = ssl.create_default_context()
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        logger.info("Email sent successfully.")
    except Exception as e:
        logger.error(f"Failed to send email: {e}")

def extract_date_from_filename(filename: str) -> str:
    """
    Extract a date in the format YYYY-MM-DD from the filename.
    If no date is found, return today's date.
    """
    match = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if match:
        return match.group(1)
    else:
        return datetime.now().strftime("%Y-%m-%d")

def paste_csv_into_excel_template(csv_path: str, template_path: str, output_path: str, sheet_name: str = "Daily Data") -> None:
    """
    Load the Excel template, read the CSV into a DataFrame, paste its contents starting at cell A1
    of the specified sheet, force a full recalculation on load (if supported), and save the result.
    """
    try:
        df = pd.read_csv(csv_path)
        wb = openpyxl.load_workbook(template_path)
        # Try to force recalculation on load (if supported)
        try:
            wb.calculation_properties.fullCalcOnLoad = True
        except AttributeError:
            logger.warning("Workbook does not support 'calculation_properties'. Consider updating openpyxl.")
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # If you don't want to remove existing formulas or formatting outside A1, do not clear the sheet
        else:
            ws = wb.create_sheet(sheet_name)
        # Write headers in row 1 starting at A1
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # Write data starting at row 2 (A2)
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        wb.save(output_path)
        logger.info(f"Pasted CSV {csv_path} into Excel template and saved as {output_path}")
    except Exception as e:
        logger.error(f"Failed to paste CSV into Excel template: {e}")

def combine_csv_files(csv_files: list[str]) -> pd.DataFrame:
    dfs = []
    for file in csv_files:
        try:
            df = pd.read_csv(file)
            dfs.append(df)
            logger.info(f"Read CSV file: {file} with {len(df)} rows.")
        except Exception as e:
            logger.error(f"Error reading {file}: {e}")
    if dfs:
        combined_df = pd.concat(dfs, ignore_index=True)
        logger.info(f"Combined CSV DataFrame has {len(combined_df)} rows.")
        return combined_df
    else:
        logger.info("No CSV files were read successfully.")
        return pd.DataFrame()

##############################################################################
#                              Main Script                                   #
##############################################################################

async def main():
    """
    Main entry point:
      1) Load creds from `creds.json`
      2) Connect to Viam (creating both data_client and cloud client)
      3) For each loc_id, fetch location name from the Fleet Management API
      4) Write that mapping (loc_id -> loc_name) to location_mapping.json (NOT in download_dir)
      5) Clear the download directory, then iterate over each location to download CSVs 
         (naming them "LocationName-OriginalFilename")
      6) For each location, choose one CSV file (the most recent) and paste its contents into an Excel template,
         saving as "Daily Report: <LocationName> - <Date>.xlsx" where <Date> is extracted from the CSV filename.
      7) Email all the daily report Excel files as attachments
    """
    logger.info("===== Starting main script =====")
    creds = load_credentials("creds.json")
    api_key = creds["API_KEY"]
    api_key_id = creds["API_KEY_ID"]
    loc_ids = creds.get("LOC_IDS", [])
    days_back = creds.get("DAYS_BACK", 1)
    smtp_host = creds["email"]["smtp_server"]
    smtp_port = creds["email"]["smtp_port"]
    sender_email = creds["email"]["from"]
    sender_password = creds["email"]["password"]
    recipient_email = creds["email"]["to"]  # Can be a list

    download_dir = "./data"
    os.makedirs(download_dir, exist_ok=True)
    clear_directory(download_dir)

    logger.info("Connecting to Viam data client ...")
    viam_client = await connect_to_viam(api_key, api_key_id)
    cloud = viam_client.app_client
    data_client = viam_client.data_client

    logger.info(f"Fetching location names for location IDs: {loc_ids}")
    loc_names = {}
    try:
        loc_names = await fetch_location_names(cloud, loc_ids)
    except Exception as e:
        logger.error(f"Could not fetch location names: {e}")

    mapping_path = "location_mapping.json"  # Outside of download_dir
    with open(mapping_path, "w") as f:
        json.dump(loc_names, f, indent=2)
    logger.info(f"Wrote location mapping to {mapping_path}")

    try:
        await download_csvs_to_directory(
            data_client, loc_ids, loc_names, days_back, download_dir
        )
    except Exception as e:
        logger.error(f"Error downloading CSVs: {e}")
    finally:
        viam_client.close()
        logger.info("Viam client closed.")

    all_files = gather_all_files_in_directory(download_dir)
    csv_files = [f for f in all_files if f.lower().endswith(".csv")]

    daily_reports = []
    template_path = "Template Sbarro Daily.xlsx"  # Excel template

    # For each location, choose the most recent CSV and paste into a daily report Excel file.
    for loc_id, loc_name in loc_names.items():
        loc_csv_files = [f for f in csv_files if os.path.basename(f).startswith(f"{loc_name}-")]
        if not loc_csv_files:
            logger.info(f"No CSV files found for location {loc_name}.")
            continue
        loc_csv_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        chosen_csv = loc_csv_files[0]
        # Extract the date from the CSV filename using a regex
        base = os.path.basename(chosen_csv)
        match = re.search(r"(\d{4}-\d{2}-\d{2})", base)
        if match:
            date_str = match.group(1)
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")
        output_excel = os.path.join(download_dir, f"Daily Report: {loc_name} - {date_str}.xlsx")
        paste_csv_into_excel_template(chosen_csv, template_path, output_excel, sheet_name="Daily Data")
        daily_reports.append(output_excel)
    
    if not daily_reports:
        logger.info("No daily report Excel files generated.")
    else:
        logger.info(f"Generated daily report Excel files: {daily_reports}")

    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    subject = f"Daily Pizza Tracking Reports from Viam For: {yesterday}"
    body = "Hello Sbarro Team! Please see attached daily report Excel files."
    if daily_reports:
        send_email_with_attachments(
            smtp_host,
            smtp_port,
            sender_email,
            sender_password,
            recipient_email,
            subject,
            body,
            daily_reports
        )
    else:
        logger.info("No daily reports to attach, email will not be sent.")

    logger.info("===== Main script finished =====")

if __name__ == "__main__":
    asyncio.run(main())
